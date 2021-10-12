import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from yattag import Doc, indent
import pandas
from xml.sax import ContentHandler, parse

import os

TRANSLATION_TEMPLATE = "%translate_sample%"
UNFINISHED_ATTRIB = {"type": "unfinished"}
VANISHED_ATTRIB = {"type": "vanished"}


class InfoContainer:
    def __init__(self):
        self.__messages = {}

    def add(self, message):
        if message not in self.__messages:
            self.__messages[message] = 1
        else:
            self.__messages[message] += 1

    def print(self):
        for message in sorted(self.__messages.keys()):
            print('    {}: {}'.format(message, self.__messages[message]))
        print('')


def consoleWriter(a, b):
    print('---------->')
    print(repr(a))
    print('    ---')
    print('    ' + repr(b))
    print('    <----------')


def xlsxToXml(filePath):

    # Load our Excel File
    wb = load_workbook(filePath + '.xlsx')
    # Getting an object of active sheet 1
    ws = wb.worksheets[0]

    # Returning returns a triplet
    doc, tag, text = Doc().tagtext()

    xml_header = '<?xml version="1.0" encoding="UTF-8" standalone = "yes"?>'
    xml_schema = '<xs:schema xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"></xs:schema>'

    doc.asis(xml_header)
    doc.asis(xml_schema)

    with tag('text'):
        for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=2):
            row = [cell.value for cell in row]
            with tag("text"):
                with tag("ru"):
                    text(row[0])
                with tag("eng"):
                    text(row[1])

    result = indent(
        doc.getvalue(),
        indentation=' ',
        indent_text=True
    )

    with open(filePath + '.xml', "w") as f:
        f.write(result)


def copyUpperCases(originalWord: str, newWord: str):
    """
    Переводчики: Использовали одно и тоже слово, а разница была в большой и маленькой буквах.
    OK = Ok ?
    """
    retVal = list(newWord)
    for i, key in enumerate(list(originalWord)):
        if i >= len(newWord):
            break

        if key.isupper():
            newKey = retVal[i].upper()
        else:
            newKey = retVal[i].lower()

        retVal[i] = newKey

    return ''.join(retVal)


def markNotTranslatedSource(needToTranslateData, source):
    """
    Отмечает не переведенный текст.
    """
    message = ET.SubElement(needToTranslateData, 'message')
    ru = ET.SubElement(message, 'ru')
    ru.text = repr(source.text)[1:-1]  # сырой текст
    eng = ET.SubElement(message, 'eng')
    eng.text = TRANSLATION_TEMPLATE
    pass


def findTranslation(updateModel, message):
    """
    Поиск перевода по заданному тексту.
    """
    source = message.find('source')
    for context in updateModel.getroot():
        sourceText = source.text
        ruText = context.find('ru').text.replace('\\n', '\n');
        engText = context.find('eng').text.replace('\\n', '\n')

        if sourceText.lower() != ruText.lower():
            continue

        # uppercases
        ruText = copyUpperCases(sourceText, ruText)
        newText = list(engText)
        newText[0] = copyUpperCases(sourceText, engText)[0]
        engText = ''.join(newText)

        returnValue = (
            ruText,
            engText
        )

        return returnValue

    return None


def findUnfinished(inputFilename, needToTranslateFilename):
    info = InfoContainer()
    needToTranslateData = ET.Element('data')
    needToTranslateModel = ET.ElementTree(needToTranslateData)

    model = ET.parse(inputFilename)
    modelRoot = model.getroot().findall('context')
    for context in modelRoot:
        messages = context.findall('message')

        for message in messages:
            translation = message.find('translation')
            isOk = True

            if translation is None:
                isOk = False

            if translation.text is None:
                isOk = False

            if translation.text == '' or translation.text == TRANSLATION_TEMPLATE:
                isOk = False

            if translation.attrib == UNFINISHED_ATTRIB:
                isOk = False

            if isOk:
                continue
            source = message.find('source')
            markNotTranslatedSource(needToTranslateData, source)
            info.add('total unfinished')

    needToTranslateModel.write(needToTranslateFilename, encoding='utf-8')
    info.print()


def isTextNull(text, ignoreSampleName=False):
    if text is None:
        return 1

    if text == '':
        return 2

    if not ignoreSampleName:
        if text == TRANSLATION_TEMPLATE:
            return 3

    return 0


def checkEachUnfinishedTranslate(inputFileName, ignoreSampleName=False):
    """
    Ищет непереведенные строки в ts файле.
    """
    print(f'Check "{inputFileName}" for unfinished translates')
    print(inputFileName)
    model = ET.parse(inputFileName)
    modelRoot = model.getroot().findall('context')
    for context in modelRoot:
        messages = context.findall('message')
        for message in messages:
            translation = message.find("translation")
            if isTranslationNull(translation, ignoreSampleName):
                translation.attrib = UNFINISHED_ATTRIB
                continue
            translation.attrib = {}
    model.write(inputFileName, encoding='utf-8')


def isTranslationNull(translation, ignoreSampleName=False):
    if translation is None:
        return 1

    if translation.text is None or translation.text == '':
        return 2

    if not ignoreSampleName:
        if translation.text == TRANSLATION_TEMPLATE:
            return 3

    return 0


def checkUnfinishedTranslate(message, text=None):
    translation = message.find("translation")
    case = isTranslationNull(translation)
    if case == 1:
        translation = ET.Element("translation")
        translation.attrib = UNFINISHED_ATTRIB
        return translation

    if case == 2 or case == 3:
        translation.attrib = UNFINISHED_ATTRIB
        return translation

    source = message.find("source")
    if text and source.lower() == text.lower():
        translation.attrib = UNFINISHED_ATTRIB
        return translation

    translation.attrib = {}
    return translation


def commitTranslation(inputFilename, updateFilename, outputFilename, forceTranslate=False, ignoreSampleName=False):
    """
    forceTranslate: перезаписывать ли уже готовые переводы
    ignoreSampleName: (не)помечать плейсхолдерное значение, как "не готовый" перевод
    """

    print(f'Start update with "{updateFilename}"')
    info = InfoContainer()
    updateModel = ET.parse(updateFilename)
    model = ET.parse(inputFilename)
    modelRoot = model.getroot().findall('context')
    for context in modelRoot:
        messages = context.findall('message')

        for message in messages:
            info.add('1. words')
            translationInfo = findTranslation(updateModel, message)
            if not translationInfo:
                info.add('2. translate not included')
                continue

            ruText = translationInfo[0]
            text = translationInfo[1]
            if isTextNull(text, ignoreSampleName):
                info.add('3. skipped')
                info.add('3.1. translate not valid')
                continue

            source = message.find('source')
            if source.text != ruText:
                info.add('3. skipped')
                info.add('3.2. translate not fits')
                continue

            translation = message.find("translation")
            if translation.attrib != UNFINISHED_ATTRIB:
                if not forceTranslate:
                    info.add('3. skipped')
                    info.add('3.3. translate is finished')
                    continue

            translation.text = text
            translation.attrib = {}
            info.add('9. updated')

    model.write(outputFilename, encoding='utf-8')
    info.print()


def clearVanished(inputFilename):
    """
    Иногда переводы могут быть помечены как Vanished в ts файле.
    Vanished - не нужны либо переводчик их не видит.
    На всякий случай?
    """
    print(f'Start clear vanished in "{inputFilename}"')
    info = InfoContainer()

    model = ET.parse(inputFilename)
    modelRoot = model.getroot().findall('context')
    for context in modelRoot:
        messagesForRemove = []
        messages = context.findall('message')
        for message in messages:
            translation = message.find('translation')
            if translation is None:
                continue

            if translation.attrib == VANISHED_ATTRIB:
                messagesForRemove.append(message)

        for message in messagesForRemove:
            context.remove(message)
    model.write(inputFilename, encoding='utf-8')


def xmlToQtTs(filePath):
    print(filePath)
    ts_original = f'{filePath}.ts'
    ts_NT = f'{filePath}_NT.ts'
    unfinishedFileName = f'{filePath}/not_translated.xml'

    if not os.path.exists(ts_original):
        with open(ts_original, 'w'): pass
    if not os.path.exists(ts_NT):
        with open(ts_NT): pass
    if not os.path.exists(unfinishedFileName):
        with open(unfinishedFileName): pass

    checkEachUnfinishedTranslate(ts_original)
    # commitTranslation(ts_original, 'update/update1.xml', ts_original, forceTranslate=True)
    # commitTranslation(ts_original, 'update/update2.xml', ts_original, forceTranslate=True)
    # commitTranslation(ts_original, 'update/update3.xml', ts_original, forceTranslate=True)
    # clearVanished(ts_original)

    # ts_NT - ТЕМПОВЫЙ ФАЙЛ ДЛЯ ДЕБАГА
    # ЕСЛИ
    # findUnfinished(ts_original, unfinishedFileName)  # если что-то не переведенно
    # commitTranslation(ts_original, unfinishedFileName, ts_NT, forceTranslate=True, ignoreSampleName=True)
    # clearVanished(ts_NT)
